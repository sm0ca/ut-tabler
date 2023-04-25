"""CSC111 FINAL PROJECT: UT-TABLER

===============================

This python module contains the functions necessary to display
the timetables to the user. This python module will take a list of
SectInfo (which represents a timetable) and display that list of
SectInfo on an Excel spreadsheet as a formatted timetable. Then,
the Excel spreadsheet will be converted to an HTML page, which will then
be displayed to the user.

Copyright and Usage Information
===============================

This file is provided solely for the users of the UT-TABLER application.
All forms of distribution of this code, whether as given or with any changes, are
expressly prohibited. For more information on copyright for CSC111 materials,
please consult our Course Syllabus.

This file is Copyright (c) 2023 Anbuselvan Ragunathan, Sanchaai Mathiyarasan, Yathusan Koneswararajah
"""
import math
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from xlsx2html import xlsx2html

from schedule_tree import SectInfo

FONT_COURSES_DAYS = Font(name='Arial', size=12, color='000000')
LST_COLORS = ['FFF1E6', 'FDE2E4', 'FAD2E1', 'E2ECE9', 'BEE1E6', 'F0EFEB', 'CDDAFD', 'DDDBCB', 'CBEEF3',
              'FFC100', 'FFEAAE', '8FBC94', 'C5E99B', 'D8DAD3', '5C9EAD']


def make_timetable(section: list[SectInfo]) -> None:
    """
    This is the most important function of this file. This function is responsible for generating the timetable and
    saving it into an Excel file in the format '.xlsx'. It then converts this file to a '.htm' file.

    Preconditions:
    - all(isinstance(item, SectInfo) for item in section)
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    days, times = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri'], ['9:00', '9:30', '10:00', '10:30', '11:00', '11:30', '12:00',
                                                        '12:30', '13:00', '13:30', '14:00', '14:30',
                                                        '15:00', '15:30', '16:00', '16:30', '17:00', '17:30', '18:00',
                                                        '18:30', '19:00', '19:30', '20:00', '20:30', '21:00']
    ws.append([''] + days)
    for i in range(0, len(times), 2):
        ws.cell(row=i + 2, column=1, value=times[i]).alignment = Alignment(horizontal='center', vertical='center',
                                                                           wrap_text=True)
        ws.cell(row=i + 2, column=10, value=times[i]).alignment = Alignment(horizontal='center', vertical='center',
                                                                            wrap_text=True)

    ws['A1'] = '1st Semester'
    ws['J1'] = '2nd Semester'
    ws['K1'] = 'Mon'
    ws['L1'] = 'Tue'
    ws['M1'] = 'Wed'
    ws['N1'] = 'Thu'
    ws['O1'] = 'Fri'
    for char in ('A', 'B', 'C', 'D', 'E', 'F', 'J', 'K', 'L', 'M', 'N', 'O'):
        ws[char + '1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    timetable_data = (times_maker(fall_winter_splitter(section, 'fall')),
                      times_maker(fall_winter_splitter(section, 'winter')))
    dict_colours = colour_assign(section)

    for day, classes in timetable_data[0].items():
        set_so_far = set()
        for time, classname in classes.items():
            if classname not in set_so_far:
                ws.cell(row=(times.index(time) + 2), column=(days.index(day) + 2),
                        value=classname).alignment = Alignment(horizontal='center',
                                                               vertical='center',
                                                               wrap_text=True)
                set_so_far.add(classname)
            fill = PatternFill(start_color=dict_colours[classname[0:10]],
                               end_color=dict_colours[classname[0:10]], fill_type='solid')
            ws.cell(row=(times.index(time) + 2), column=(days.index(day) + 2)).fill = fill

    for day, classes in timetable_data[1].items():
        set_so_far = set()
        for time, classname in classes.items():
            if classname not in set_so_far:
                ws.cell(row=(times.index(time) + 2), column=(days.index(day) + 11),
                        value=classname).alignment = Alignment(horizontal='center',
                                                               vertical='center',
                                                               wrap_text=True)
                set_so_far.add(classname)
            fill = PatternFill(start_color=dict_colours[classname[0:10]],
                               end_color=dict_colours[classname[0:10]], fill_type='solid')
            ws.cell(row=(times.index(time) + 2), column=(days.index(day) + 11)).fill = fill

    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'J', 'K', 'L', 'M', 'N', 'O'):
        ws.column_dimensions[letter].width = 14
        for num in range(1, 27):
            ws[letter + str(num)].font = FONT_COURSES_DAYS

    ws.column_dimensions['A'].width, ws.column_dimensions['J'].width = 17, 17
    ws['A1'].font, ws['J1'].font = [Font(name='Arial', size=12, color='000000', bold=True)] * 2
    wb.save('timetable.xlsx')
    xlsx2html('timetable.xlsx', 'timetable.htm')


def times_maker(section_information: list[SectInfo]) -> dict[str, dict[str, str]]:
    """
    This function takes in a list of section information and returns a dictionary with a key
    value of the day of the week and a value dictionary that has its key value as the time and its
    value as the course name

    Preconditions:
    - section_information != []
    - all(isinstance(item, SectInfo) for item in section)
    """
    dict_final = {'Mon': {}, 'Tue': {}, 'Wed': {}, 'Thu': {}, 'Fri': {}}
    for section in section_information:
        course = section.course
        sect = section.sect
        for instance in section.room_times:
            day = instance
            if not section.room_times[day]:
                time_day = 'NA'
            else:
                time_day = start_and_end(section.room_times[day][0][1])
            for lst in section.room_times[instance]:
                interval = times_maker_helper(lst[1])
                for time in interval:
                    dict_final[day][time] = course + '\n' + section.instructors + '\n' + sect + '\n' + time_day
    return dict_final


def times_maker_helper(interval: tuple) -> list[str]:
    """
    Helper method for the times_maker function. This takes a time interval and returns a list of strings that associates
    to all half-hour points of that time interval.

    Preconditions:
    - interval != ()
    """
    lst_times = []
    for j in range(math.floor(interval[0]), (math.floor(interval[1]) + 1)):
        if ('.0' in str(interval[1])) and (j == math.floor(interval[1])):
            lst_times.append(str(j) + ':00')
        else:
            lst_times.append(str(j) + ':00')
            lst_times.append(str(j) + ':30')
    lst_times.pop()
    return lst_times


def start_and_end(interval: tuple[float, float]) -> str:
    """
    Return start and end time of the time interval

    Preconditions:
    - interval != ()
    """
    str_so_far = ''
    t1, t2 = interval
    t1_floor = math.floor(t1)
    t2_floor = math.floor(t2)
    if '.5' in str(t1):
        str_so_far += str(t1_floor) + ':30 - '
    else:
        str_so_far += str(t1_floor) + ':00 - '

    if '.5' in str(t2):
        str_so_far += str(t2_floor) + ':30'
    else:
        str_so_far += str(t2_floor) + ':00'
    return str_so_far


def colour_assign(lec_info: list[SectInfo]) -> dict[str, str]:
    """
    Return a dictionary with the courses assigned to a colour

    Preconditions:
    - lec_info != []
    - all(isinstance(item, SectInfo) for item in section)
    """
    dict_colour_course = {}
    set_colours_used = set()
    lst_courses = []
    for lec in lec_info:
        lst_courses.append(lec.course)
    for i in range(len(lst_courses)):
        if len(set_colours_used) == 15:
            set_colours_used.clear()
        dict_colour_course[lst_courses[i]] = LST_COLORS[i]
        set_colours_used.add(LST_COLORS[i])
    return dict_colour_course


def fall_winter_splitter(sec_info: list[SectInfo], semester: str) -> list[SectInfo]:
    """
    This function takes in a list of section information and returns a list of courses that are
    specified by <semester>

    Preconditions:
    - all(isinstance(item, SectInfo) for item in section)
    - semester in {'fall', 'winter'}
    """
    sem_courses = []
    if semester == 'fall':
        for info in sec_info:
            if ('-F' in info.course) or ('-Y' in info.course):
                sem_courses.append(info)
    else:
        for info in sec_info:
            if ('-S' in info.course) or ('-Y' in info.course):
                sem_courses.append(info)
    return sem_courses


if __name__ == '__main__':
    import doctest
    import python_ta

    doctest.testmod(verbose=True)
    python_ta.check_all(config={
        'max-line-length': 120,
        'extra-imports': ['math', 'openpyxl', 'schedule_tree', 'openpyxl.styles', 'xlsx2html'],
        'disable': ['too-many-branches', 'too-many-nested-blocks', 'too-many-locals'],
        'allowed-io': ['Alignment', 'PatternFill', 'Font', 'SectInfo', 'xlsx2html']

    })
